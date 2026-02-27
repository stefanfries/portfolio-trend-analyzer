"""Save analysis results to Excel/CSV file for historical comparison."""

import csv
import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import PatternFill

from app.settings import ResultsSaverSettings

# Excel formatting constants
EXECUTE_MARKER = "✅ EXECUTE NOW"
WAITING_MARKER = "⏳"
COLOR_EXECUTE_GREEN = "C6EFCE"  # Light green for execute now
COLOR_WAITING_YELLOW = "FFEB9C"  # Light yellow for pending


def save_results_to_xlsx(results: list[dict], depot_name: str = "mega_trend_folger"):
    """
    Save analysis results to an Excel file with proper formatting for German locale.

    Args:
        results: List of result dictionaries from the main analysis
        depot_name: Name of the depot being analyzed

    Returns:
        Path to the saved file
    """
    settings = ResultsSaverSettings()
    
    if not results:
        print("⚠️ No results to save.")
        return

    # Create results directory if it doesn't exist
    results_dir = Path(settings.results_dir)
    results_dir.mkdir(exist_ok=True)

    # Generate filename with timestamp
    timestamp = datetime.now().strftime(settings.results_timestamp_format)
    filename = results_dir / f"{depot_name}_{timestamp}.xlsx"

    # Process results to extract numeric values
    processed_results = []
    for result in results:
        processed = result.copy()

        # Extract numeric value from "Current Price" (e.g., "2.25 €" -> 2.25)
        if "Current Price" in processed:
            price_str = processed["Current Price"]
            match = re.search(r"([\d.]+)", price_str)
            if match:
                processed["Current Price"] = float(match.group(1))

        # Extract numeric value from "Drawdown %" (e.g., "-50.44" -> -50.44)
        if "Drawdown %" in processed:
            drawdown_str = str(processed["Drawdown %"])
            match = re.search(r"(-?[\d.]+)", drawdown_str)
            if match:
                processed["Drawdown %"] = float(match.group(1)) / 100  # Convert to decimal

        # Convert ADX to float if it's a string
        if "ADX" in processed:
            adx_str = str(processed["ADX"])
            match = re.search(r"([\d.]+)", adx_str)
            if match:
                processed["ADX"] = float(match.group(1))

        processed_results.append(processed)

    # Create DataFrame
    df = pd.DataFrame(processed_results)

    # Reorder columns for better readability
    column_order = [
        "WKN",
        "Name",
        "Current Price",
        "Drawdown %",
        "ADX",
        "Supertrend",
        "Trend Signal",
        "Execution Recommendation",
        "Reason",
    ]
    # Add any other columns that might exist
    for col in df.columns:
        if col not in column_order:
            column_order.append(col)
    # Only use columns that exist
    column_order = [col for col in column_order if col in df.columns]
    df = df[column_order]

    # Write to Excel with formatting
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=settings.excel_sheet_name, index=False)

        # Get the worksheet
        worksheet = writer.sheets[settings.excel_sheet_name]

        # Format columns
        for idx, col in enumerate(df.columns, start=1):
            col_letter = chr(64 + idx)  # A, B, C, etc.

            if col == "Current Price":
                # Format as currency with 2 decimals and Euro symbol
                for row in range(2, len(df) + 2):  # Start from row 2 (after header)
                    cell = worksheet[f"{col_letter}{row}"]
                    cell.number_format = '#,##0.00 "€"'

            elif col == "Drawdown %":
                # Format as percentage with 2 decimals
                for row in range(2, len(df) + 2):
                    cell = worksheet[f"{col_letter}{row}"]
                    cell.number_format = "0.00%"

            elif col == "ADX":
                # Format as number with 1 decimal
                for row in range(2, len(df) + 2):
                    cell = worksheet[f"{col_letter}{row}"]
                    cell.number_format = "0.0"

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, settings.excel_max_column_width)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Enable auto-filter on header row
        worksheet.auto_filter.ref = worksheet.dimensions

        # Apply background colors based on Execution Recommendation
        if "Execution Recommendation" in df.columns:
            exec_col_idx = df.columns.get_loc("Execution Recommendation") + 1
            exec_col_letter = chr(64 + exec_col_idx)

            # Define colors using constants
            green_fill = PatternFill(
                start_color=COLOR_EXECUTE_GREEN, end_color=COLOR_EXECUTE_GREEN, fill_type="solid"
            )
            yellow_fill = PatternFill(
                start_color=COLOR_WAITING_YELLOW, end_color=COLOR_WAITING_YELLOW, fill_type="solid"
            )

            for row in range(2, len(df) + 2):  # Start from row 2 (after header)
                cell_value = worksheet[f"{exec_col_letter}{row}"].value
                if cell_value and EXECUTE_MARKER in str(cell_value):
                    # Green for execute now
                    for col_idx in range(1, len(df.columns) + 1):
                        worksheet[f"{chr(64 + col_idx)}{row}"].fill = green_fill
                elif cell_value and WAITING_MARKER in str(cell_value):
                    # Yellow for waiting/pending
                    for col_idx in range(1, len(df.columns) + 1):
                        worksheet[f"{chr(64 + col_idx)}{row}"].fill = yellow_fill

        # Freeze header row
        worksheet.freeze_panes = "A2"

    print(f"\n💾 Results saved to: {filename}")
    print(f"   Securities: {len(results)}")
    print(f"   Timestamp: {timestamp}")
    print("   Format: Excel (.xlsx) with proper number formatting")

    return filename


def save_results_to_csv(results: list[dict], depot_name: str = "mega_trend_folger"):
    """
    Save analysis results to a CSV file with timestamp.

    Args:
        results: List of result dictionaries from the main analysis
        depot_name: Name of the depot being analyzed
    """
    settings = ResultsSaverSettings()
    
    if not results:
        print("⚠️ No results to save.")
        return

    # Create results directory if it doesn't exist
    results_dir = Path(settings.results_dir)
    results_dir.mkdir(exist_ok=True)

    # Generate filename with timestamp
    timestamp = datetime.now().strftime(settings.results_timestamp_format)
    filename = results_dir / f"{depot_name}_{timestamp}.csv"

    # Get all possible keys from results
    all_keys = set()
    for result in results:
        all_keys.update(result.keys())

    # Sort keys for consistent column order
    fieldnames = sorted(all_keys)

    # Write to CSV
    with open(filename, "w", newline="", encoding="utf-8") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(results)

    print(f"\n💾 Results saved to: {filename}")
    print(f"   Securities: {len(results)}")
    print(f"   Timestamp: {timestamp}")

    return filename


def load_results_from_csv(filename: str) -> list[dict]:
    """
    Load previously saved results from CSV file.

    Args:
        filename: Path to the CSV file

    Returns:
        List of result dictionaries
    """
    results = []
    with open(filename, encoding="utf-8") as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            results.append(dict(row))

    return results


def list_available_results(depot_name: str = None) -> list[Path]:
    """
    List all available result files.

    Args:
        depot_name: Optional depot name to filter by

    Returns:
        List of Path objects for result files
    """
    settings = ResultsSaverSettings()
    results_dir = Path(settings.results_dir)
    if not results_dir.exists():
        return []

    if depot_name:
        pattern = f"{depot_name}_*.csv"
    else:
        pattern = "*.csv"

    return sorted(results_dir.glob(pattern), reverse=True)
